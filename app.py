from flask import Flask, render_template, request
import os
import openpyxl


app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html', message='')

@app.route('/grades', methods=['POST'])
def grades():
    sitting_num_input = request.form['sitting_num']
    college = request.form['college']

    try:
        sitting_num = int(sitting_num_input)
    except ValueError:
        return render_template('index.html', message='Sitting number must be a number')

    excel_path = os.path.join(os.path.dirname(__file__), 'grades', f'{college}.xlsx')

    try:
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook[workbook.sheetnames[0]]

        excel_data = list(sheet.iter_rows(values_only=True))  # Load the data as a list of rows

        matching_row = None

        for row in excel_data:
            if str(sitting_num) in str(row[1]):
                matching_row = row
                break

        if matching_row:
            seventh_row_data = excel_data[6]  # Get the 7th row (index 6 in Python)
            co =college
            return render_template('college.html', row=matching_row, college_name=college, seventh_row_data=seventh_row_data,co =co)

    except Exception as error:
        print('Error reading Excel file:', error)

    return render_template('index.html', message='Sitting number or college not found')

if __name__ == '__main__':
        app.run(host='0.0.0.0', port=5000)

